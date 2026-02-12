#!/usr/bin/env python3
"""
Convert vehicle data (CSV or Excel) to JSON format for the Isuzu Motors website.

Usage:
    python excel_to_json.py              # Read from templates/ CSV files
    python excel_to_json.py vehicles.xlsx  # Read from Excel file

Requirements:
    pip install openpyxl

This script reads vehicle data and generates data/vehicles.json
"""

import json
import csv
import sys
from pathlib import Path
from openpyxl import load_workbook
import re

# Mapping of CSV filenames/Excel sheet names to category data.
# This is defined once and reused by both CSV and Excel processing functions.
CATEGORY_MAPPING = {
    "Passenger Vehicles": {
        "csv_name": "Passenger Vehicles.csv",
        "id": "passenger",
        "name": "Passenger Vehicles",
        "subtitle": "Personal & Family",
        "icon": "👨‍👩‍👧‍👦",
        "description": "Personal & Family vehicles for everyday comfort"
    },
    "Light Commercial": {
        "csv_name": "Light Commercial.csv",
        "id": "light-commercial",
        "name": "Light Commercial Vehicles",
        "subtitle": "Small Business",
        "icon": "🏪",
        "description": "Small Business solutions for efficient operations"
    },
    "Medium Duty Trucks": {
        "csv_name": "Medium Duty Trucks.csv",
        "id": "medium-duty",
        "name": "Light to Medium Duty Trucks",
        "subtitle": "N-Series & F-Series",
        "icon": "🚛",
        "description": "N-Series & F-Series for versatile hauling"
    },
    "Heavy Duty GIGA": {
        "csv_name": "Heavy Duty GIGA.csv",
        "id": "heavy-duty",
        "name": "Heavy Duty & Special Purpose",
        "subtitle": "GIGA",
        "icon": "💪",
        "description": "GIGA series for demanding operations"
    }
}

def csv_to_json():
    """Convert CSV files from templates/ folder to JSON format."""
    
    templates_dir = Path("templates")
    if not templates_dir.exists():
        print(f"Error: {templates_dir}/ directory not found.")
        sys.exit(1)

    categories = []

    for _, category_info in CATEGORY_MAPPING.items():
        csv_name = category_info["csv_name"]
        csv_path = templates_dir / csv_name
        if not csv_path.exists():
            print(f"Warning: File '{csv_path}' not found. Skipping...")
            continue

        models = []

        # Read CSV file (use utf-8-sig to strip BOM if present)
        with open(csv_path, 'r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row is None or not row.get('Model'):
                    break

                raw_price = row.get('2026 Price Range (SRP)', '')
                models.append({
                    "model": row.get('Model', '').strip(),
                    "variant": row.get('Variant', '').strip(),
                    "priceRange": format_price_range(raw_price)
                })

        category = {
            "id": category_info["id"],
            "name": category_info["name"],
            "subtitle": category_info["subtitle"],
            "icon": category_info["icon"],
            "description": category_info["description"],
            "models": models
        }
        categories.append(category)

    # Create and save JSON
    output_data = {"categories": categories}
    save_json(output_data)


def excel_to_json(excel_file):
    """Convert Excel workbook to JSON format matching the website structure."""
    
    if not Path(excel_file).exists():
        print(f"Error: File '{excel_file}' not found.")
        sys.exit(1)

    wb = load_workbook(excel_file)
    categories = []

    for sheet_name, category_info in CATEGORY_MAPPING.items():
        if sheet_name not in wb.sheetnames:
            print(f"Warning: Sheet '{sheet_name}' not found. Skipping...")
            continue

        ws = wb[sheet_name]
        models = []

        # Read data starting from row 2 (row 1 is headers)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:  # Stop at empty rows
                break
            
            model = row[0] if row[0] else ""
            variant = row[1] if row[1] else ""
            price_range = row[2] if row[2] else ""

            if model:  # Only add if model name is not empty
                models.append({
                    "model": str(model).strip(),
                    "variant": str(variant).strip(),
                    "priceRange": format_price_range(price_range)
                })

        category = {
            "id": category_info["id"],
            "name": category_info["name"],
            "subtitle": category_info["subtitle"],
            "icon": category_info["icon"],
            "description": category_info["description"],
            "models": models
        }
        categories.append(category)

    # Create and save JSON
    output_data = {"categories": categories}
    save_json(output_data)


def save_json(data):
    """Save data to JSON file."""
    output_file = Path("data/vehicles.json")
    output_file.parent.mkdir(parents=True, exist_ok=True)
    
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"✓ Successfully generated '{output_file}'")
    print(f"  Total categories: {len(data['categories'])}")
    for cat in data['categories']:
        print(f"  - {cat['name']}: {len(cat['models'])} models")


def format_price_range(raw):
    """Normalize and format price ranges.

    Examples:
      P1070000 - P1140000  -> ₱1,070,000 – ₱1,140,000
      1070000-1140000      -> ₱1,070,000 – ₱1,140,000
      Price on Request      -> Price on Request (unchanged)
    """
    if raw is None:
        return ""

    s = str(raw).strip()
    if not s:
        return ""

    # If it's already clearly non-numeric (e.g., Price on Request), return as-is
    if 'request' in s.lower():
        return "Price on Request"

    # Split ranges on hyphen or dash
    parts = re.split(r"\s*[–-]\s*", s)
    formatted_parts = []

    for part in parts:
        # Extract numeric and decimal characters
        num_raw = re.sub(r"[^0-9.]", "", part)
        if not num_raw:
            # can't parse number, return original whole string
            return s

        try:
            if '.' in num_raw:
                value = float(num_raw)
                # Format with two decimals if needed
                formatted = f"₱{value:,.2f}"
            else:
                value = int(num_raw)
                formatted = f"₱{value:,}"
        except Exception:
            return s

        formatted_parts.append(formatted)

    if len(formatted_parts) == 1:
        return formatted_parts[0]
    else:
        return ' – '.join(formatted_parts)


if __name__ == "__main__":
    if len(sys.argv) > 1:
        # Excel file provided
        excel_to_json(sys.argv[1])
    else:
        # Read from CSV templates
        csv_to_json()
